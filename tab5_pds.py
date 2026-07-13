# FILE NAME: tab5_pds.py

import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.drawing.image import Image as xlImage

def render(c, t2, t3):
    st.markdown("### 📄 Process Data Sheet (PDS) Generator")
    st.write("Upload your standard Excel (`.xlsx`) template. The system will automatically map the calculated dimensions, populate the nozzle schedule, and insert the GA drawings into your file.")
    
    st.info("""
    **💡 How to prepare your Excel Template:**
    Type these exact tags anywhere in your Excel file. The system will replace them with calculated values:
    * `{{TAG}}` ➔ Equipment Tag No.
    * `{{FLUID}}` ➔ Handled Fluid
    * `{{ROOF}}` ➔ Roof Type
    * `{{DIA}}` ➔ Tank Diameter (mm)
    * `{{HEIGHT}}` ➔ Total Shell Height (mm)
    * `{{VOL}}` ➔ Working Volume (m³)
    * `{{HHLL}}`, `{{HLL}}`, `{{LLL}}`, `{{LLLL}}` ➔ Elevation Levels (mm)
    * `{{NOZZLES}}` ➔ System will replace this row with the full Nozzle Schedule Table.
    * `{{DRAWING_GA}}` ➔ System will replace this tag with the Side View Image.
    * `{{DRAWING_TOP}}` ➔ System will replace this tag with the Top View Image.
    """)

    uploaded_template = st.file_uploader("Browse & Upload your Excel Template (.xlsx)", type=["xlsx"], key="pds_uploader")

    if uploaded_template is not None:
        if st.button("🔄 Generate & Download PDS", type="primary"):
            with st.spinner("Processing your PDS... Please wait."):
                try:
                    # 1. Prepare mapped data
                    mapped_data = {
                        "{{TAG}}": str(c['tag_no']),
                        "{{FLUID}}": str(c['fluid']),
                        "{{ROOF}}": str(c['roof_type']),
                        "{{DIA}}": str(int(t2['calc_diameter_m'] * 1000)),
                        "{{HEIGHT}}": str(int(t2['final_top_mm'])),
                        "{{VOL}}": str(round(t2['final_volume_m3'], 2)),
                        "{{HHLL}}": str(int(t2['final_hhll_mm'])),
                        "{{HLL}}": str(int(t2['final_hll_mm'])),
                        "{{LLL}}": str(int(t2['final_lll_mm'])),
                        "{{LLLL}}": str(int(t2['final_llll_mm']))
                    }

                    # 2. Convert Plotly Figures to Images in memory
                    side_img_bytes = t3['fig_side'].to_image(format="png", width=800, height=800)
                    top_img_bytes = t3['fig_polar'].to_image(format="png", width=800, height=800)
                    
                    xl_side_img = xlImage(io.BytesIO(side_img_bytes))
                    xl_top_img = xlImage(io.BytesIO(top_img_bytes))
                    
                    # 3. Load Excel Workbook
                    wb = openpyxl.load_workbook(uploaded_template)
                    ws = wb.active # Assuming editing the first sheet
                    
                    nozzle_start_row = None
                    nozzle_start_col = None

                    # 4. Search and Replace text tags, and locate special anchors
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                text = cell.value
                                
                                # Replace basic text placeholders
                                for key, val in mapped_data.items():
                                    if key in text:
                                        text = text.replace(key, val)
                                        cell.value = text
                                
                                # Find Nozzle Table Anchor
                                if "{{NOZZLES}}" in text:
                                    nozzle_start_row = cell.row
                                    nozzle_start_col = cell.column
                                    cell.value = "" # Clear the tag
                                    
                                # Find Drawing Anchors
                                if "{{DRAWING_GA}}" in text:
                                    cell.value = ""
                                    ws.add_image(xl_side_img, cell.coordinate)
                                    
                                if "{{DRAWING_TOP}}" in text:
                                    cell.value = ""
                                    ws.add_image(xl_top_img, cell.coordinate)

                    # 5. Insert Nozzle Table if anchor was found
                    if nozzle_start_row is not None:
                        # Write Headers
                        headers = ["Mark", "Position", "Service", "Size (mm)", "Rating", "Internals", "Remarks"]
                        for c_idx, header in enumerate(headers):
                            ws.cell(row=nozzle_start_row, column=nozzle_start_col + c_idx, value=header)
                            
                        # Write Data
                        nozzle_df = t3['nozzles_df']
                        curr_row = nozzle_start_row + 1
                        for _, n_row in nozzle_df.iterrows():
                            ws.cell(row=curr_row, column=nozzle_start_col, value=str(n_row['Mark']))
                            ws.cell(row=curr_row, column=nozzle_start_col+1, value=str(n_row['Position']))
                            ws.cell(row=curr_row, column=nozzle_start_col+2, value=str(n_row['Service']))
                            ws.cell(row=curr_row, column=nozzle_start_col+3, value=str(n_row['Size (mm)']))
                            ws.cell(row=curr_row, column=nozzle_start_col+4, value=str(n_row['Flange Rating']))
                            ws.cell(row=curr_row, column=nozzle_start_col+5, value=str(n_row['Internals']))
                            ws.cell(row=curr_row, column=nozzle_start_col+6, value=str(n_row.get('Remarks', '')))
                            curr_row += 1

                    # 6. Save modified workbook to memory
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    # 7. Provide Download Button
                    st.success("✅ PDS Generated Successfully!")
                    st.download_button(
                        label="📥 Download Filled PDS (Excel)",
                        data=output,
                        file_name=f"{c['tag_no']}_PDS.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except Exception as e:
                    st.error(f"❌ An error occurred during generation: {e}")
                    st.error("Tip: Ensure you have `openpyxl` and `kaleido` installed via pip for Excel and Image processing.")